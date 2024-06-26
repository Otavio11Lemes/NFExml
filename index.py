import xml.etree.ElementTree as ET
from openpyxl import Workbook

# Carregar o arquivo XML
tree = ET.parse('./Xmls/52240602924249000119550020018676171243672280-nfe.xml')
root = tree.getroot()

# Definir o namespace
ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

# Criar um novo arquivo XLSX
wb = Workbook()
ws_emitente = wb.create_sheet(title="Emitente")
ws_imposto = wb.create_sheet(title="Imposto")
ws_produto = wb.create_sheet(title="Produto")

# Definir os cabeçalhos
headers_emitente = ["CNPJ", "inscricao_estadual", "razao_social", "logradouro", "numero", "bairro", "municipio", "uf"]
headers_imposto = ["aliquota_ICMS", "aliquota_ICMS_ST", "aliquota_IPI"]
headers_produto = ["cEAN", "descricao", "ncm", "cst", "cfop", "unidade", "quantidade", "valor_unitario"]

# Preencher o cabeçalho na primeira linha
ws_emitente.append(headers_emitente)
ws_imposto.append(headers_imposto)
ws_produto.append(headers_produto)

# Extrair dados do emitente
emitente = root.find('.//nfe:emit', ns)
CNPJ = emitente.find('nfe:CNPJ', ns).text
inscricao_estadual = emitente.find('nfe:IE', ns).text
razao_social = emitente.find('nfe:xNome', ns).text
logradouro = emitente.find('nfe:enderEmit/nfe:xLgr', ns).text
numero = emitente.find('nfe:enderEmit/nfe:nro', ns).text
bairro = emitente.find('nfe:enderEmit/nfe:xBairro', ns).text
municipio = emitente.find('nfe:enderEmit/nfe:xMun', ns).text
uf = emitente.find('nfe:enderEmit/nfe:UF', ns).text

# Adicionar os dados à próxima linha na aba Emitente
ws_emitente.append([CNPJ, inscricao_estadual, razao_social, logradouro, numero, bairro, municipio, uf])

# Extrair dados dos produtos e impostos
for det in root.findall('.//nfe:det', ns):
    # Produto
    cEAN = det.find('nfe:prod/nfe:cEAN', ns).text
    descricao = det.find('nfe:prod/nfe:xProd', ns).text
    ncm = det.find('nfe:prod/nfe:NCM', ns).text
    cst = det.find('nfe:imposto/nfe:ICMS//nfe:CST', ns).text
    cfop = det.find('nfe:prod/nfe:CFOP', ns).text
    unidade = det.find('nfe:prod/nfe:uCom', ns).text
    quantidade = det.find('nfe:prod/nfe:qCom', ns).text
    valor_unitario = det.find('nfe:prod/nfe:vUnCom', ns).text

    # Adicionar os dados do produto à próxima linha na aba Produto
    ws_produto.append([cEAN, descricao, ncm, cst, cfop, unidade, quantidade, valor_unitario])

    # Impostos
    aliquota_ICMS = det.find('nfe:imposto/nfe:ICMS//nfe:pICMS', ns)
    aliquota_ICMS_ST = det.find('nfe:imposto/nfe:ICMS//nfe:pICMSST', ns)
    aliquota_IPI = det.find('nfe:imposto/nfe:IPI//nfe:pIPI', ns)

    # Verificar se os elementos são None antes de acessar o atributo text
    aliquota_ICMS = aliquota_ICMS.text if aliquota_ICMS is not None else None
    aliquota_ICMS_ST = aliquota_ICMS_ST.text if aliquota_ICMS_ST is not None else None
    aliquota_IPI = aliquota_IPI.text if aliquota_IPI is not None else None

    # Adicionar os dados do imposto à próxima linha na aba Imposto
    ws_imposto.append([aliquota_ICMS, aliquota_ICMS_ST, aliquota_IPI])

# Salvar o arquivo
wb.save("dados_nfe.xlsx")
